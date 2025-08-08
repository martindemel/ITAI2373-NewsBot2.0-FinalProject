#!/usr/bin/env python3
"""
Export Manager for NewsBot 2.0
Advanced export capabilities for analysis results and reports
"""

import numpy as np
import pandas as pd
from typing import Dict, List, Any, Optional, Union
import logging
from datetime import datetime
import json
import csv
import pickle
import os
import zipfile
import tempfile
from pathlib import Path

# Document generation libraries
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    logging.warning("reportlab not available for PDF generation")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logging.warning("openpyxl not available for Excel export")

class ExportManager:
    """
    Comprehensive export manager for NewsBot 2.0 analysis results
    """
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialize export manager
        
        Args:
            config: Configuration dictionary
        """
        self.config = config or {}
        
        # Export formats and their handlers
        self.export_handlers = {
            'json': self._export_json,
            'csv': self._export_csv,
            'excel': self._export_excel,
            'pdf': self._export_pdf,
            'html': self._export_html,
            'xml': self._export_xml,
            'yaml': self._export_yaml,
            'pickle': self._export_pickle,
            'zip': self._export_zip_archive
        }
        
        # Default export settings
        self.default_settings = {
            'include_metadata': True,
            'include_timestamps': True,
            'compress_large_files': True,
            'max_file_size_mb': 100,
            'encoding': 'utf-8',
            'date_format': '%Y-%m-%d %H:%M:%S'
        }
        
        # Export templates
        self.export_templates = {
            'analysis_report': {
                'sections': ['executive_summary', 'methodology', 'results', 'insights', 'appendix'],
                'formats': ['pdf', 'html', 'excel']
            },
            'data_export': {
                'sections': ['raw_data', 'processed_data', 'metadata'],
                'formats': ['csv', 'excel', 'json']
            },
            'visualization_package': {
                'sections': ['charts', 'interactive_plots', 'static_images'],
                'formats': ['html', 'zip']
            },
            'api_response': {
                'sections': ['results', 'metadata', 'statistics'],
                'formats': ['json', 'xml']
            }
        }
        
        # File size and performance tracking
        self.export_stats = {
            'total_exports': 0,
            'exports_by_format': {},
            'avg_export_time': 0,
            'total_data_exported_mb': 0
        }
    
    def export_analysis_results(self, results: Dict[str, Any], 
                               export_format: str = 'json',
                               output_path: Optional[str] = None,
                               template: str = 'analysis_report',
                               settings: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        Export comprehensive analysis results
        
        Args:
            results: Analysis results to export
            export_format: Target export format
            output_path: Output file path (auto-generated if None)
            template: Export template to use
            settings: Additional export settings
            
        Returns:
            Export operation results
        """
        start_time = datetime.now()
        
        # Merge settings
        export_settings = {**self.default_settings, **(settings or {})}
        
        # Validate format
        if export_format not in self.export_handlers:
            raise ValueError(f"Unsupported export format: {export_format}")
        
        # Generate output path if not provided
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"newsbot_analysis_{timestamp}.{export_format}"
            output_path = os.path.join(self.config.get('export_dir', 'exports'), filename)
        
        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        export_result = {
            'export_id': f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            'timestamp': start_time.isoformat(),
            'format': export_format,
            'template': template,
            'output_path': output_path,
            'status': 'processing'
        }
        
        try:
            # Prepare data for export
            export_data = self._prepare_export_data(results, template, export_settings)
            
            # Execute export
            handler = self.export_handlers[export_format]
            export_details = handler(export_data, output_path, export_settings)
            
            export_result.update(export_details)
            export_result['status'] = 'completed'
            
            # Calculate export time
            end_time = datetime.now()
            export_time = (end_time - start_time).total_seconds()
            export_result['export_time'] = export_time
            
            # Update statistics
            self._update_export_stats(export_format, export_time, export_result.get('file_size', 0))
            
            logging.info(f"Export completed: {output_path}")
            
        except Exception as e:
            logging.error(f"Export failed: {e}")
            export_result['status'] = 'failed'
            export_result['error'] = str(e)
        
        return export_result
    
    def export_dashboard_data(self, dashboard_data: Dict[str, Any],
                            export_format: str = 'html',
                            output_path: Optional[str] = None) -> Dict[str, Any]:
        """
        Export dashboard data with visualizations
        
        Args:
            dashboard_data: Dashboard data including charts and metrics
            export_format: Export format (html, pdf, zip)
            output_path: Output file path
            
        Returns:
            Export results
        """
        logging.info(f"Exporting dashboard data to {export_format}...")
        
        if export_format == 'html':
            return self._export_dashboard_html(dashboard_data, output_path)
        elif export_format == 'pdf':
            return self._export_dashboard_pdf(dashboard_data, output_path)
        elif export_format == 'zip':
            return self._export_dashboard_zip(dashboard_data, output_path)
        else:
            raise ValueError(f"Unsupported dashboard export format: {export_format}")
    
    def export_bulk_data(self, data_collections: Dict[str, Any],
                        export_format: str = 'zip',
                        output_path: Optional[str] = None) -> Dict[str, Any]:
        """
        Export multiple data collections in bulk
        
        Args:
            data_collections: Dictionary of data collections to export
            export_format: Export format (typically zip for bulk)
            output_path: Output file path
            
        Returns:
            Bulk export results
        """
        logging.info(f"Exporting bulk data ({len(data_collections)} collections)...")
        
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"newsbot_bulk_export_{timestamp}.{export_format}"
        
        export_result = {
            'collections_exported': len(data_collections),
            'export_format': export_format,
            'output_path': output_path,
            'individual_exports': {},
            'timestamp': datetime.now().isoformat()
        }
        
        try:
            if export_format == 'zip':
                # Create zip archive with multiple exports
                with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for collection_name, collection_data in data_collections.items():
                        # Export each collection to temporary file
                        temp_file = self._export_collection_to_temp(collection_name, collection_data)
                        
                        # Add to zip
                        arcname = f"{collection_name}/{os.path.basename(temp_file)}"
                        zipf.write(temp_file, arcname)
                        
                        export_result['individual_exports'][collection_name] = arcname
                        
                        # Clean up temp file
                        os.unlink(temp_file)
                
                export_result['status'] = 'completed'
                export_result['file_size'] = os.path.getsize(output_path)
            
            else:
                raise ValueError(f"Bulk export format {export_format} not supported")
                
        except Exception as e:
            logging.error(f"Bulk export failed: {e}")
            export_result['status'] = 'failed'
            export_result['error'] = str(e)
        
        return export_result
    
    def create_analysis_report(self, analysis_data: Dict[str, Any],
                             report_type: str = 'comprehensive',
                             output_format: str = 'pdf',
                             output_path: Optional[str] = None) -> Dict[str, Any]:
        """
        Create comprehensive analysis report
        
        Args:
            analysis_data: Complete analysis data
            report_type: Type of report (comprehensive, executive, technical)
            output_format: Output format (pdf, html, docx)
            output_path: Output file path
            
        Returns:
            Report generation results
        """
        logging.info(f"Creating {report_type} analysis report in {output_format} format...")
        
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"newsbot_{report_type}_report_{timestamp}.{output_format}"
        
        report_result = {
            'report_type': report_type,
            'output_format': output_format,
            'output_path': output_path,
            'timestamp': datetime.now().isoformat()
        }
        
        try:
            # Generate report content based on type
            report_content = self._generate_report_content(analysis_data, report_type)
            
            # Export in requested format
            if output_format == 'pdf':
                self._create_pdf_report(report_content, output_path)
            elif output_format == 'html':
                self._create_html_report(report_content, output_path)
            else:
                raise ValueError(f"Report format {output_format} not supported")
            
            report_result['status'] = 'completed'
            report_result['file_size'] = os.path.getsize(output_path)
            
        except Exception as e:
            logging.error(f"Report creation failed: {e}")
            report_result['status'] = 'failed'
            report_result['error'] = str(e)
        
        return report_result
    
    # Individual export format handlers
    
    def _export_json(self, data: Dict[str, Any], output_path: str, 
                    settings: Dict[str, Any]) -> Dict[str, Any]:
        """Export data as JSON"""
        
        try:
            with open(output_path, 'w', encoding=settings['encoding']) as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
            
            file_size = os.path.getsize(output_path)
            
            return {
                'file_size': file_size,
                'encoding': settings['encoding'],
                'records_exported': self._count_records(data)
            }
            
        except Exception as e:
            raise Exception(f"JSON export failed: {e}")
    
    def _export_csv(self, data: Dict[str, Any], output_path: str,
                   settings: Dict[str, Any]) -> Dict[str, Any]:
        """Export data as CSV"""
        
        try:
            # Convert data to DataFrame if possible
            if isinstance(data, dict) and 'data' in data:
                df_data = data['data']
            else:
                df_data = data
            
            if isinstance(df_data, pd.DataFrame):
                df = df_data
            elif isinstance(df_data, list) and df_data:
                df = pd.DataFrame(df_data)
            elif isinstance(df_data, dict):
                # Try to create DataFrame from dict
                try:
                    df = pd.DataFrame.from_dict(df_data, orient='index')
                except:
                    # Fallback: create single-row DataFrame
                    df = pd.DataFrame([df_data])
            else:
                raise ValueError("Data cannot be converted to CSV format")
            
            # Export to CSV
            df.to_csv(output_path, index=False, encoding=settings['encoding'])
            
            file_size = os.path.getsize(output_path)
            
            return {
                'file_size': file_size,
                'records_exported': len(df),
                'columns_exported': len(df.columns),
                'encoding': settings['encoding']
            }
            
        except Exception as e:
            raise Exception(f"CSV export failed: {e}")
    
    def _export_excel(self, data: Dict[str, Any], output_path: str,
                     settings: Dict[str, Any]) -> Dict[str, Any]:
        """Export data as Excel workbook"""
        
        if not HAS_OPENPYXL:
            raise Exception("openpyxl library required for Excel export")
        
        try:
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            sheets_created = 0
            
            # Handle different data structures
            if isinstance(data, dict):
                for sheet_name, sheet_data in data.items():
                    if sheet_name.startswith('_'):  # Skip metadata
                        continue
                    
                    # Create sheet
                    ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
                    
                    if isinstance(sheet_data, pd.DataFrame):
                        self._write_dataframe_to_sheet(ws, sheet_data)
                    elif isinstance(sheet_data, list):
                        self._write_list_to_sheet(ws, sheet_data)
                    elif isinstance(sheet_data, dict):
                        self._write_dict_to_sheet(ws, sheet_data)
                    
                    sheets_created += 1
            
            # If no sheets created, create a summary sheet
            if sheets_created == 0:
                ws = wb.create_sheet(title="Data")
                self._write_dict_to_sheet(ws, data)
                sheets_created = 1
            
            # Save workbook
            wb.save(output_path)
            
            file_size = os.path.getsize(output_path)
            
            return {
                'file_size': file_size,
                'sheets_created': sheets_created,
                'records_exported': self._count_records(data)
            }
            
        except Exception as e:
            raise Exception(f"Excel export failed: {e}")
    
    def _export_pdf(self, data: Dict[str, Any], output_path: str,
                   settings: Dict[str, Any]) -> Dict[str, Any]:
        """Export data as PDF report"""
        
        if not HAS_REPORTLAB:
            raise Exception("reportlab library required for PDF export")
        
        try:
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            story.append(Paragraph("NewsBot 2.0 Analysis Report", title_style))
            story.append(Spacer(1, 20))
            
            # Add timestamp
            timestamp = datetime.now().strftime(settings['date_format'])
            story.append(Paragraph(f"Generated on: {timestamp}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Add data sections
            self._add_data_to_pdf_story(story, data, styles)
            
            # Build PDF
            doc.build(story)
            
            file_size = os.path.getsize(output_path)
            
            return {
                'file_size': file_size,
                'pages_generated': 'variable',  # Would need page counting
                'sections_included': len(data) if isinstance(data, dict) else 1
            }
            
        except Exception as e:
            raise Exception(f"PDF export failed: {e}")
    
    def _export_html(self, data: Dict[str, Any], output_path: str,
                    settings: Dict[str, Any]) -> Dict[str, Any]:
        """Export data as HTML report"""
        
        try:
            html_content = self._generate_html_content(data, settings)
            
            with open(output_path, 'w', encoding=settings['encoding']) as f:
                f.write(html_content)
            
            file_size = os.path.getsize(output_path)
            
            return {
                'file_size': file_size,
                'encoding': settings['encoding'],
                'sections_included': len(data) if isinstance(data, dict) else 1
            }
            
        except Exception as e:
            raise Exception(f"HTML export failed: {e}")
    
    def _export_xml(self, data: Dict[str, Any], output_path: str,
                   settings: Dict[str, Any]) -> Dict[str, Any]:
        """Export data as XML"""
        
        try:
            xml_content = self._dict_to_xml(data, 'NewsBot_Analysis')
            
            with open(output_path, 'w', encoding=settings['encoding']) as f:
                f.write(xml_content)
            
            file_size = os.path.getsize(output_path)
            
            return {
                'file_size': file_size,
                'encoding': settings['encoding'],
                'records_exported': self._count_records(data)
            }
            
        except Exception as e:
            raise Exception(f"XML export failed: {e}")
    
    def _export_yaml(self, data: Dict[str, Any], output_path: str,
                    settings: Dict[str, Any]) -> Dict[str, Any]:
        """Export data as YAML"""
        
        try:
            import yaml
            
            with open(output_path, 'w', encoding=settings['encoding']) as f:
                yaml.dump(data, f, default_flow_style=False, allow_unicode=True)
            
            file_size = os.path.getsize(output_path)
            
            return {
                'file_size': file_size,
                'encoding': settings['encoding'],
                'records_exported': self._count_records(data)
            }
            
        except ImportError:
            raise Exception("PyYAML library required for YAML export")
        except Exception as e:
            raise Exception(f"YAML export failed: {e}")
    
    def _export_pickle(self, data: Dict[str, Any], output_path: str,
                      settings: Dict[str, Any]) -> Dict[str, Any]:
        """Export data as pickle file"""
        
        try:
            with open(output_path, 'wb') as f:
                pickle.dump(data, f, protocol=pickle.HIGHEST_PROTOCOL)
            
            file_size = os.path.getsize(output_path)
            
            return {
                'file_size': file_size,
                'pickle_protocol': pickle.HIGHEST_PROTOCOL,
                'records_exported': self._count_records(data)
            }
            
        except Exception as e:
            raise Exception(f"Pickle export failed: {e}")
    
    def _export_zip_archive(self, data: Dict[str, Any], output_path: str,
                           settings: Dict[str, Any]) -> Dict[str, Any]:
        """Export data as compressed zip archive"""
        
        try:
            files_in_archive = 0
            
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Export main data as JSON
                main_data_json = json.dumps(data, indent=2, default=str)
                zipf.writestr('analysis_data.json', main_data_json)
                files_in_archive += 1
                
                # Export metadata
                metadata = {
                    'export_timestamp': datetime.now().isoformat(),
                    'newsbot_version': '2.0',
                    'data_summary': self._generate_data_summary(data)
                }
                metadata_json = json.dumps(metadata, indent=2, default=str)
                zipf.writestr('metadata.json', metadata_json)
                files_in_archive += 1
                
                # If data contains DataFrames, export as CSV too
                if isinstance(data, dict):
                    for key, value in data.items():
                        if isinstance(value, pd.DataFrame):
                            csv_data = value.to_csv(index=False)
                            zipf.writestr(f'{key}.csv', csv_data)
                            files_in_archive += 1
            
            file_size = os.path.getsize(output_path)
            
            return {
                'file_size': file_size,
                'files_in_archive': files_in_archive,
                'compression_ratio': self._calculate_compression_ratio(data, file_size)
            }
            
        except Exception as e:
            raise Exception(f"ZIP export failed: {e}")
    
    # Helper methods
    
    def _prepare_export_data(self, results: Dict[str, Any], template: str,
                           settings: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare data for export based on template"""
        
        template_config = self.export_templates.get(template, {})
        sections = template_config.get('sections', ['results'])
        
        export_data = {}
        
        # Add metadata if requested
        if settings.get('include_metadata', True):
            export_data['_metadata'] = {
                'export_timestamp': datetime.now().isoformat(),
                'template_used': template,
                'newsbot_version': '2.0',
                'export_settings': settings
            }
        
        # Add sections based on template
        for section in sections:
            if section in results:
                export_data[section] = results[section]
            elif section == 'executive_summary':
                export_data[section] = self._generate_executive_summary(results)
            elif section == 'methodology':
                export_data[section] = self._generate_methodology_section(results)
            elif section == 'insights':
                export_data[section] = self._extract_insights(results)
            elif section == 'appendix':
                export_data[section] = self._generate_appendix(results)
        
        return export_data
    
    def _generate_executive_summary(self, results: Dict[str, Any]) -> Dict[str, Any]:
        """Generate executive summary from results"""
        
        summary = {
            'total_articles_analyzed': 0,
            'analysis_timeframe': 'unknown',
            'key_findings': [],
            'performance_metrics': {},
            'recommendations': []
        }
        
        # Extract key metrics
        if 'classification' in results:
            class_results = results['classification']
            if 'total_articles' in class_results:
                summary['total_articles_analyzed'] = class_results['total_articles']
        
        if 'sentiment' in results:
            sentiment_results = results['sentiment']
            if 'sentiment_distribution' in sentiment_results:
                dist = sentiment_results['sentiment_distribution']
                dominant_sentiment = max(dist, key=dist.get)
                summary['key_findings'].append(f"Dominant sentiment: {dominant_sentiment} ({dist[dominant_sentiment]:.1%})")
        
        if 'topics' in results:
            topic_results = results['topics']
            if 'num_topics' in topic_results:
                summary['key_findings'].append(f"Identified {topic_results['num_topics']} main topics")
        
        return summary
    
    def _generate_methodology_section(self, results: Dict[str, Any]) -> Dict[str, Any]:
        """Generate methodology section"""
        
        methodology = {
            'data_sources': 'News articles dataset',
            'preprocessing_steps': [
                'Text cleaning and normalization',
                'Tokenization and stop word removal',
                'Feature extraction (TF-IDF, embeddings)'
            ],
            'analysis_methods': [],
            'evaluation_metrics': []
        }
        
        # Determine which analyses were performed
        if 'classification' in results:
            methodology['analysis_methods'].append('Machine Learning Classification')
            methodology['evaluation_metrics'].extend(['Accuracy', 'Precision', 'Recall', 'F1-Score'])
        
        if 'sentiment' in results:
            methodology['analysis_methods'].append('Sentiment Analysis')
            methodology['evaluation_metrics'].append('Sentiment Distribution')
        
        if 'topics' in results:
            methodology['analysis_methods'].append('Topic Modeling (LDA/NMF)')
            methodology['evaluation_metrics'].append('Topic Coherence')
        
        if 'entities' in results:
            methodology['analysis_methods'].append('Named Entity Recognition')
            methodology['evaluation_metrics'].append('Entity Extraction Coverage')
        
        return methodology
    
    def _extract_insights(self, results: Dict[str, Any]) -> List[str]:
        """Extract key insights from results"""
        
        insights = []
        
        # Classification insights
        if 'classification' in results:
            class_results = results['classification']
            if 'accuracy' in class_results:
                accuracy = class_results['accuracy']
                insights.append(f"Classification model achieved {accuracy:.1%} accuracy")
        
        # Sentiment insights
        if 'sentiment' in results:
            sentiment_results = results['sentiment']
            if 'correlation_vader_textblob' in sentiment_results:
                correlation = sentiment_results['correlation_vader_textblob']
                insights.append(f"Sentiment analysis methods show {correlation:.3f} correlation")
        
        # Topic insights
        if 'topics' in results:
            topic_results = results['topics']
            if 'coherence_score' in topic_results:
                coherence = topic_results['coherence_score']
                insights.append(f"Topic model coherence score: {coherence:.3f}")
        
        return insights
    
    def _generate_appendix(self, results: Dict[str, Any]) -> Dict[str, Any]:
        """Generate appendix with technical details"""
        
        appendix = {
            'technical_specifications': {
                'python_version': '3.8+',
                'key_libraries': ['scikit-learn', 'pandas', 'numpy', 'nltk', 'spacy'],
                'model_parameters': {}
            },
            'data_statistics': {},
            'performance_benchmarks': {}
        }
        
        # Add model parameters if available
        for analysis_type in ['classification', 'sentiment', 'topics']:
            if analysis_type in results:
                analysis_results = results[analysis_type]
                if 'model_parameters' in analysis_results:
                    appendix['technical_specifications']['model_parameters'][analysis_type] = analysis_results['model_parameters']
        
        return appendix
    
    def _write_dataframe_to_sheet(self, worksheet, df: pd.DataFrame):
        """Write DataFrame to Excel worksheet"""
        
        # Write headers
        for col_idx, column_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=column_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # Write data
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _write_list_to_sheet(self, worksheet, data_list: List[Any]):
        """Write list data to Excel worksheet"""
        
        if not data_list:
            return
        
        # If list contains dictionaries, treat as table
        if isinstance(data_list[0], dict):
            # Get all unique keys for headers
            all_keys = set()
            for item in data_list:
                if isinstance(item, dict):
                    all_keys.update(item.keys())
            
            headers = list(all_keys)
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
            
            # Write data
            for row_idx, item in enumerate(data_list, 2):
                if isinstance(item, dict):
                    for col_idx, header in enumerate(headers, 1):
                        value = item.get(header, '')
                        worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        else:
            # Simple list - write as single column
            worksheet.cell(row=1, column=1, value='Data').font = Font(bold=True)
            for row_idx, item in enumerate(data_list, 2):
                worksheet.cell(row=row_idx, column=1, value=item)
    
    def _write_dict_to_sheet(self, worksheet, data_dict: Dict[str, Any]):
        """Write dictionary data to Excel worksheet"""
        
        # Write as key-value pairs
        worksheet.cell(row=1, column=1, value='Key').font = Font(bold=True)
        worksheet.cell(row=1, column=2, value='Value').font = Font(bold=True)
        
        for row_idx, (key, value) in enumerate(data_dict.items(), 2):
            worksheet.cell(row=row_idx, column=1, value=key)
            
            # Convert complex values to string
            if isinstance(value, (dict, list)):
                value_str = json.dumps(value, indent=2, default=str)
            else:
                value_str = str(value)
            
            worksheet.cell(row=row_idx, column=2, value=value_str)
    
    def _add_data_to_pdf_story(self, story: List, data: Dict[str, Any], styles):
        """Add data sections to PDF story"""
        
        for section_name, section_data in data.items():
            if section_name.startswith('_'):  # Skip metadata sections
                continue
            
            # Section header
            story.append(Paragraph(section_name.replace('_', ' ').title(), styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Section content
            if isinstance(section_data, dict):
                # Create table from dictionary
                table_data = [['Key', 'Value']]
                for key, value in section_data.items():
                    table_data.append([key, str(value)[:100]])  # Truncate long values
                
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)
            
            elif isinstance(section_data, list):
                # Create bulleted list
                for item in section_data[:10]:  # Limit to first 10 items
                    story.append(Paragraph(f"• {str(item)}", styles['Normal']))
            
            else:
                # Plain text
                story.append(Paragraph(str(section_data), styles['Normal']))
            
            story.append(Spacer(1, 20))
    
    def _generate_html_content(self, data: Dict[str, Any], settings: Dict[str, Any]) -> str:
        """Generate HTML content from data"""
        
        html_parts = [
            "<!DOCTYPE html>",
            "<html>",
            "<head>",
            "<title>NewsBot 2.0 Analysis Report</title>",
            "<meta charset='utf-8'>",
            "<style>",
            "body { font-family: Arial, sans-serif; margin: 40px; }",
            "h1 { color: #2c3e50; border-bottom: 2px solid #3498db; }",
            "h2 { color: #34495e; margin-top: 30px; }",
            "table { border-collapse: collapse; width: 100%; margin: 20px 0; }",
            "th, td { border: 1px solid #ddd; padding: 12px; text-align: left; }",
            "th { background-color: #f2f2f2; font-weight: bold; }",
            "tr:nth-child(even) { background-color: #f9f9f9; }",
            ".metadata { background-color: #ecf0f1; padding: 15px; margin: 20px 0; border-radius: 5px; }",
            "</style>",
            "</head>",
            "<body>",
            "<h1>NewsBot 2.0 Analysis Report</h1>"
        ]
        
        # Add timestamp
        timestamp = datetime.now().strftime(settings['date_format'])
        html_parts.append(f"<div class='metadata'><strong>Generated:</strong> {timestamp}</div>")
        
        # Add data sections
        for section_name, section_data in data.items():
            if section_name.startswith('_'):  # Skip metadata sections
                continue
            
            html_parts.append(f"<h2>{section_name.replace('_', ' ').title()}</h2>")
            
            if isinstance(section_data, dict):
                html_parts.append("<table>")
                html_parts.append("<tr><th>Key</th><th>Value</th></tr>")
                for key, value in section_data.items():
                    html_parts.append(f"<tr><td>{key}</td><td>{str(value)}</td></tr>")
                html_parts.append("</table>")
            
            elif isinstance(section_data, list):
                html_parts.append("<ul>")
                for item in section_data:
                    html_parts.append(f"<li>{str(item)}</li>")
                html_parts.append("</ul>")
            
            else:
                html_parts.append(f"<p>{str(section_data)}</p>")
        
        html_parts.extend(["</body>", "</html>"])
        
        return "\n".join(html_parts)
    
    def _dict_to_xml(self, data: Dict[str, Any], root_name: str) -> str:
        """Convert dictionary to XML format"""
        
        def dict_to_xml_recursive(d, root):
            xml_str = f"<{root}>"
            
            for key, value in d.items():
                # Sanitize key for XML
                clean_key = re.sub(r'[^a-zA-Z0-9_]', '_', str(key))
                
                if isinstance(value, dict):
                    xml_str += dict_to_xml_recursive(value, clean_key)
                elif isinstance(value, list):
                    for item in value:
                        if isinstance(item, dict):
                            xml_str += dict_to_xml_recursive(item, clean_key)
                        else:
                            xml_str += f"<{clean_key}>{str(item)}</{clean_key}>"
                else:
                    xml_str += f"<{clean_key}>{str(value)}</{clean_key}>"
            
            xml_str += f"</{root}>"
            return xml_str
        
        import re
        
        xml_header = '<?xml version="1.0" encoding="UTF-8"?>\n'
        xml_content = dict_to_xml_recursive(data, root_name)
        
        return xml_header + xml_content
    
    def _count_records(self, data: Any) -> int:
        """Count records in data structure"""
        
        if isinstance(data, pd.DataFrame):
            return len(data)
        elif isinstance(data, list):
            return len(data)
        elif isinstance(data, dict):
            total = 0
            for value in data.values():
                if isinstance(value, (list, pd.DataFrame)):
                    total += self._count_records(value)
                elif isinstance(value, dict):
                    total += len(value)
            return total if total > 0 else 1
        else:
            return 1
    
    def _calculate_compression_ratio(self, original_data: Dict[str, Any], 
                                   compressed_size: int) -> float:
        """Calculate compression ratio"""
        
        try:
            # Estimate original size
            original_json = json.dumps(original_data, default=str)
            original_size = len(original_json.encode('utf-8'))
            
            if original_size > 0:
                return compressed_size / original_size
            else:
                return 1.0
        except:
            return 1.0
    
    def _generate_data_summary(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate summary of data for metadata"""
        
        summary = {
            'total_sections': len(data),
            'data_types': {},
            'record_counts': {},
            'size_estimate': 0
        }
        
        for key, value in data.items():
            summary['data_types'][key] = type(value).__name__
            summary['record_counts'][key] = self._count_records(value)
        
        # Estimate size
        try:
            json_str = json.dumps(data, default=str)
            summary['size_estimate'] = len(json_str.encode('utf-8'))
        except:
            summary['size_estimate'] = 0
        
        return summary
    
    def _export_collection_to_temp(self, collection_name: str, 
                                  collection_data: Any) -> str:
        """Export collection to temporary file"""
        
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"{collection_name}.json")
        
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(collection_data, f, indent=2, default=str)
        
        return temp_file
    
    def _update_export_stats(self, export_format: str, export_time: float, file_size: int):
        """Update export statistics"""
        
        self.export_stats['total_exports'] += 1
        
        if export_format not in self.export_stats['exports_by_format']:
            self.export_stats['exports_by_format'][export_format] = 0
        self.export_stats['exports_by_format'][export_format] += 1
        
        # Update average export time
        total = self.export_stats['total_exports']
        current_avg = self.export_stats['avg_export_time']
        self.export_stats['avg_export_time'] = ((current_avg * (total - 1)) + export_time) / total
        
        # Update total data exported
        self.export_stats['total_data_exported_mb'] += file_size / (1024 * 1024)
    
    def get_export_statistics(self) -> Dict[str, Any]:
        """Get export statistics"""
        return self.export_stats.copy()
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported export formats"""
        return list(self.export_handlers.keys())
    
    def get_export_templates(self) -> Dict[str, Any]:
        """Get available export templates"""
        return self.export_templates.copy()